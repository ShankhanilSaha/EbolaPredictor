import openpyxl
import numpy as np
import matplotlib.pyplot as plt

# Load workbook and worksheet
wb = openpyxl.load_workbook('train_data.xlsx')
ws = wb.active

lat = []
long = []
deaths = []

# Iterate through rows, starting from the second row
for i in range(2, ws.max_row + 1):
    if ws.cell(row=i, column=1).value is not None and \
       ws.cell(row=i, column=2).value is not None and \
       ws.cell(row=i, column=3).value is not None:
        lat.append(float(ws.cell(row=i, column=1).value))
        long.append(float(ws.cell(row=i, column=2).value))
        deaths.append(float(ws.cell(row=i, column=3).value))

wb.close()

# Convert to numpy arrays
x = np.array(lat)
y = np.array(long)
z = np.array(deaths)

# Plotting
fig = plt.figure()
ax = plt.axes(projection='3d')
ax.grid()

ax.scatter(x, y, z,s=5)
ax.set_title('Deaths Graph From available info')

# Set axes labels
ax.set_xlabel('Latitude', labelpad=20)
ax.set_ylabel('Longitude', labelpad=20)
ax.set_zlabel('Deaths', labelpad=20)
plt.show()